"""Excel report generation utilities for the POSE ecosystem."""

from __future__ import annotations

import io
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models import PoseKeypoint, PoseSession


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="2563EB")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_SUBHEADER_FILL = PatternFill("solid", fgColor="DBEAFE")
_SUBHEADER_FONT = Font(bold=True, color="1E3A8A")


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def _style_header_row(ws, row_num: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# Report builders
# ---------------------------------------------------------------------------

def build_sessions_report(db: Session, user_id: int | None = None) -> bytes:
    """
    Generate an Excel workbook with two sheets:
    - Summary: one row per session
    - Keypoints: all keypoint measurements

    Args:
        db: SQLAlchemy session.
        user_id: If provided, filter sessions to this user; otherwise include all.

    Returns:
        Raw bytes of the .xlsx file.
    """
    query = db.query(PoseSession)
    if user_id is not None:
        query = query.filter(PoseSession.user_id == user_id)
    sessions: list[PoseSession] = query.order_by(PoseSession.recorded_at.desc()).all()

    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------
    # Sheet 1 – Summary
    # ------------------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Sessions"

    summary_headers = ["ID", "User ID", "Label", "Recorded At", "Duration (s)", "Keypoints Count", "Notes"]
    ws_summary.append(summary_headers)
    _style_header_row(ws_summary, 1, len(summary_headers))

    for s in sessions:
        ws_summary.append([
            s.id,
            s.user_id,
            s.session_label,
            s.recorded_at.isoformat() if s.recorded_at else None,
            s.duration_seconds,
            len(s.keypoints),
            s.notes,
        ])

    _auto_width(ws_summary)

    # ------------------------------------------------------------------
    # Sheet 2 – Keypoints
    # ------------------------------------------------------------------
    ws_kp = wb.create_sheet("Keypoints")

    kp_headers = ["Session ID", "Keypoint", "X", "Y", "Z", "Confidence"]
    ws_kp.append(kp_headers)
    _style_header_row(ws_kp, 1, len(kp_headers))

    for s in sessions:
        for kp in s.keypoints:
            ws_kp.append([s.id, kp.keypoint_name, kp.x, kp.y, kp.z, kp.confidence])

    _auto_width(ws_kp)

    # ------------------------------------------------------------------
    # Metadata sheet
    # ------------------------------------------------------------------
    ws_meta = wb.create_sheet("Report Info")
    ws_meta.append(["Generated At", datetime.now(timezone.utc).isoformat()])
    ws_meta.append(["Total Sessions", len(sessions)])
    ws_meta.append(["Filter User ID", user_id if user_id is not None else "All"])
    _auto_width(ws_meta)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
